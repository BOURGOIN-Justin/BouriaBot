import discord
from discord.ext import commands
from discord import app_commands
from nbt import nbt
from collections import Counter
import json
import os
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import Reference, ProjectedPieChart, PieChart
import math


def generer_excel_farm(nom_schematic, inventaire_blocs, dico_noms, dico_prix):
    """
    Permet de générer un document excel répertoriant le nom des blocs, la quantité et le prix associé
    Args:
        nom_schematic (str) : Le nom du fichier en schematic.
        inventaire_blocs (dict) : Dictionnaire qui contient les IDs et une quantité associée.
        dico_noms (dict) : Dictionnaire qui contient les IDs et un nom associé.
        dico_prix(dict) : Dictionnaire qui contient le nom des blocs et un prix associé.
    Returns:
        str : Nom du fichier excel généré.

    """
    # On crée le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste de Farm"

    # On cache le quadrillage gris par défaut
    ws.sheet_view.showGridLines = False

    # 1. On écrit l'en-tête
    headers = ["Nom du Bloc", "ID", "Quantité", "Prix Total ($)"]
    ws.append(headers)

    # 2. On remplit les données ligne par ligne
    for id_bloc, quantite in inventaire_blocs.items():
        nom = dico_noms.get(id_bloc, "Inconnu")
        prix_unitaire = dico_prix.get(nom, 0)
        total_prix = quantite * prix_unitaire
        ws.append([nom, id_bloc, quantite, total_prix])

    # 3. Création du tableau dynamique
    nb_lignes = ws.max_row
    zone_tableau = f"A1:D{nb_lignes}"

    tab = Table(displayName="TableauDeFarm", ref=zone_tableau)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # On crée un trait fin de la même couleur bleue que ton ancien en-tête
    trait_bleu = Side(border_style="thin", color="366092")
    bordure_bleue = Border(top=trait_bleu, left=trait_bleu, right=trait_bleu, bottom=trait_bleu)

    # On applique cette bordure sur absolument TOUTES les cases de notre tableau
    for row in ws.iter_rows(min_row=1, max_row=nb_lignes, min_col=1, max_col=4):
        for cell in row:
            cell.border = bordure_bleue

    police_en_tete = Font(bold=True, color="000000")
    for cell in ws[1]:
        cell.font = police_en_tete

    # 4. On règle la largeur des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    projected_pie = ProjectedPieChart()
    projected_pie.type = "pie"
    projected_pie.splitType = "val"
    nbGratuit = 0
    nbPayant= 0



    for id_bloc, quantite in inventaire_blocs.items():
        nom = dico_noms.get(id_bloc, "Inconnu")
        prix_unitaire = dico_prix.get(nom,0)
        if prix_unitaire == 0:
            # indice 0 pour le prix à zero
            nbGratuit += 1
        else:
            # indice 1 pour un prix normal
            nbPayant += 1

    font = Font(name="Cambria", color="000000", size=11, bold=True)
    border = Border(left=Side(border_style="thin",
                          color='366092'),
                right=Side(border_style="thin",
                           color='366092'),
                top=Side(border_style="thin",
                         color='366092'),
                bottom=Side(border_style="thin",
                            color='366092'),
                diagonal=Side(border_style="thin",
                              color='366092'),
                diagonal_direction=0,
                vertical=Side(border_style="thin",
                              color='366092'),
                horizontal=Side(border_style="thin",
                               color='366092')
               )
    ws["G1"] = "Type"
    ws['G1'].font = font
    ws['G1'].border = border
    ws["H1"] = "Nombre"
    ws['H1'].font = font
    ws['H1'].border = border
    ws["G2"] = "Gratuit"
    ws['G2'].border = border
    ws["H2"] = nbGratuit
    ws['H2'].border = border
    ws["G3"] = "Payant"
    ws['G3'].border = border
    ws["H3"] = nbPayant
    ws['H3'].border = border

    # Création du graphique
    chart = PieChart()
    chart.title = "Répartition des blocs payants et gratuits"

    # 1. Références des données (Colonne H : Nombres) et catégories (Colonne G : Noms)
    data = Reference(ws, min_col=8, min_row=1, max_row=3)  # min_row=1 car titles_from_data=True
    cats = Reference(ws, min_col=7, min_row=2, max_row=3)  # Les étiquettes "Gratuit" / "Payant"

    # 3. Liaison des données au graphique
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # 4. Placement sur la feuille
    ws.add_chart(chart, "G4")

    # 5. On sauvegarde avec un nom propre
    nom_fichier = f"Liste_Farm_{nom_schematic}.xlsx"
    wb.save(nom_fichier)
    return nom_fichier

def lecture_schematic(name, dicoB, dicoP):
    """
    Permet de lire un fichier au format schematic
    Args:
        name (str) : Nom du fichier à analyser
        dicoB (dict) : Dictionnaire qui contient les IDs de blocs et le nom associé
        dicoP (dict) : Dictionnaire qui contient le nom de blocs et leur prix associé

    Returns:
        float : Prix total du schematic
        list : Liste des blocs manquants
        int : Nombre total de blocs du schematic
        dict : Dictionnaire qui contient les IDs de blocs et une quantité associé
    """
    nbtfile = nbt.NBTFile(name, 'rb')
    print("Clés du fichier :", nbtfile.keys())

    blocks = nbtfile['Blocks'].value
    data = nbtfile['Data'].value
    add_blocks = nbtfile['AddBlocks'].value if 'AddBlocks' in nbtfile else None

    liste_ids_complets = []

    for i in range(len(blocks)):
        vrai_id = blocks[i]

        if add_blocks:
            add_val = (add_blocks[i // 2] >> (0 if i % 2 else 4)) & 0x0F
            vrai_id = (add_val << 8) + vrai_id

        if vrai_id != 0:
            cle_format = f"{vrai_id}:{data[i]}"
            liste_ids_complets.append(cle_format)

    count = Counter(liste_ids_complets)

    prix_total = 0
    blocs_manquants = []

    for cle_recherche, quantite in count.items():
        nom_du_bloc = dicoB.get(cle_recherche, "Bloc Inconnu")
        if nom_du_bloc == "Bloc Inconnu":
            print(f"ID inexistant : {cle_recherche}")
            if cle_recherche not in blocs_manquants:
                blocs_manquants.append(cle_recherche)
            with open('all.json', 'w', encoding='utf-8') as fichier:
                json.dump(dicoB, fichier, indent=4, ensure_ascii=False)
        prix_bloc = dicoP.get(nom_du_bloc, 0)
        sous_total = prix_bloc * quantite
        prix_total += sous_total
        print(f"ID Minecraft [{cle_recherche}] : {quantite} blocs | name : {nom_du_bloc}")
    print("-" * 40)
    print("Nombre de blocs au total:", sum(count.values()))
    print("Prix du schematic: ", prix_total)
    return prix_total, blocs_manquants, sum(count.values()), count

def lecture_fichier_json(name):
    """
    Permet de lire un fichier au format json pour récupérer les éléments des blocks
    Args:
        name (str) : Nom du fichier au format json à lire
    """
    with open(name, 'r', encoding='utf-8') as fichier:
        donnees = json.load(fichier)
    donnees_lisibles = {}
    for bloc in donnees:
        data_val = bloc.get("Data", 0)
        cle = f"{bloc['block_id']}:{data_val}"
        donnees_lisibles[cle] = bloc["display_name_fr"]
    print("Données chargées: ", len(donnees_lisibles))
    return donnees_lisibles

def lire_json(nom_fichier):
    """
    Lit un fichier JSON et renvoie son contenu brut.
    Si le fichier n'existe pas, renvoie un dictionnaire vide.
    """
    if os.path.exists(nom_fichier):
        with open(nom_fichier, 'r', encoding='utf-8') as fichier:
            return json.load(fichier)
    return {}

BIBLIO_WE = lire_json("commandes.json")
BIBLIO_CRAFT = lire_json("craft.json")

def recherche_craft(item: str, quantite, dictionnaire_recette: dict):
    quantite = float(quantite)
    panier_total = Counter()

    if item in dictionnaire_recette:

        for ingredient, quantite_ingredient in dictionnaire_recette[item].items():

            if ingredient == "image":
                continue
            quantite_total = quantite * quantite_ingredient
            sous_ingredient = recherche_craft(ingredient, quantite_total, dictionnaire_recette)
            panier_total.update(sous_ingredient)
    else:
        panier_total[item] += quantite
    return panier_total


class NGBuildBot(commands.Bot):
    def __init__(self):
        # Configuration des intents
        intents = discord.Intents.default()
        intents.message_content = True
        # On initialise avec un préfixe bidon car on utilise les Slash Commands
        super().__init__(command_prefix="!", intents=intents)
    # Cette fonction synchronise les commandes "/" avec Discord au démarrage
    async def setup_hook(self):
        await self.tree.sync()
        print("🔄 Commandes Slash synchronisées avec succès !")
bot = NGBuildBot()


@bot.event
async def on_ready():
    print(f'✅ Connecté en tant que {bot.user} !')
    print('Le bot est prêt à recevoir des commandes.')

@bot.tree.command(name="ping", description="Vérifie si le bot est en ligne")
async def ping(interaction: discord.Interaction):
    await interaction.response.send_message("Pong ! Je suis bien en ligne et prêt à analyser tes schematics !")


class BoutonExcel(discord.ui.View):
    def __init__(self, nom_propre, count, donnees_totale, donnees_prix):
        super().__init__(timeout=None)  # Le bouton ne se désactive jamais
        # On sauvegarde les données en mémoire pour quand le joueur cliquera
        self.nom_propre = nom_propre
        self.count = count
        self.donnees_totale = donnees_totale
        self.donnees_prix = donnees_prix

    @discord.ui.button(label="📊 Télécharger l'Excel", style=discord.ButtonStyle.green)
    async def telecharger(self, interaction: discord.Interaction, button: discord.ui.Button):
        # On indique qu'on prépare le fichier (en mode invisible pour les autres)
        await interaction.response.defer(ephemeral=True)

        # On génère le fichier UNIQUEMENT au moment du clic
        nom_temp = f"temp_{interaction.id}"
        generation_excel = generer_excel_farm(nom_temp, self.count, self.donnees_totale, self.donnees_prix)
        with open(generation_excel, 'rb') as f:
            discord_file = discord.File(f, filename=f"Farm_{self.nom_propre}.xlsx")
            await interaction.followup.send(content="Voici ton fichier :", file=discord_file)

        # On supprime le fichier après l'envoi
        if os.path.exists(generation_excel):
            os.remove(generation_excel)


@bot.tree.command(name="devis", description="Génère un devis à partir d'un fichier .schematic")
@app_commands.describe(fichier="Le fichier .schematic à analyser")
async def devis(interaction: discord.Interaction, fichier: discord.Attachment):

    if not fichier.filename.endswith(".schematic"):

        await interaction.response.send_message("❌ Erreur : Le fichier doit être un `.schematic` !", ephemeral=True)
        return

    await interaction.response.defer()
    try:

        with open("all.json", 'r', encoding='utf-8') as fichiertot:
            donnees_totale = json.load(fichiertot)
        with open("prix.json", 'r', encoding='utf-8') as f_prix:
            donnees_prix = json.load(f_prix)

        chemin_local = fichier.filename
        await fichier.save(chemin_local)

        prix, liste_erreurs, blocks_total, count = lecture_schematic(chemin_local, donnees_totale, donnees_prix)
        nom_propre = fichier.filename.replace('.schematic', '')
        couleur_facture = discord.Color.red() if len(liste_erreurs) > 0 else discord.Color.green()

        facture = discord.Embed(
            title=f"<:Groupfqsf:1431359830676996167>  Devis pour le schematic : **{nom_propre}**",
            description="Voici l'analyse des coûts. Les blocs obtenables gratuitement sont comptabilisés à 0$ (Marbre, minerais, etc).",
            color=couleur_facture
        )
        facture.set_thumbnail(url=bot.user.display_avatar.url)
        facture.add_field(name="<:28993mcgrassblock:1406936610746007622> Nombre de blocs", value=f"**{blocks_total}**", inline=True)
        facture.add_field(name="<:itemmoney:1403438157596196906> Prix total", value=f"**{prix} $**", inline=False)
        facture.add_field(name="", value="", inline=False)

        # Gestion des blocs inconnus
        if len(liste_erreurs) > 0:
            premiers_ids = liste_erreurs[:10]
            texte_erreur = "\n 🔸ID " + "\n 🔸ID ".join(premiers_ids)
            message_erreur = f"Voici les premiers : {texte_erreur} \n ..."
            facture.add_field(
                name=f"<:8056engagedinsuspectedspamactiv:1406904855536074823>  {len(liste_erreurs)} blocs inconnus (comptés à 0$)",
                value=message_erreur, inline=False)

        view_bouton = BoutonExcel(nom_propre, count, donnees_totale, donnees_prix)
        await interaction.followup.send(embed=facture, view=view_bouton)

    except Exception as e:
        await interaction.followup.send(f"⚠️ Une erreur inattendue est survenue : {e}")
    finally:
        if os.path.exists(chemin_local):
            os.remove(chemin_local)

async def auto_type_cmd(interaction: discord.Interaction,current: str) -> list[app_commands.Choice[str]]:
    # Propose les catégories principales
    choix = list(BIBLIO_WE.keys())
    return [app_commands.Choice(name=c.capitalize(), value=c) for c in choix if current.lower() in c.lower()][:25]


async def auto_nom(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
    # Récupère ce que le joueur a choisi dans la case précédente
    type_choisi = interaction.namespace.type_commande

    if type_choisi in BIBLIO_WE:
        choix = list(BIBLIO_WE[type_choisi].keys())
        return [app_commands.Choice(name=c.capitalize(), value=c) for c in choix if current.lower() in c.lower()][:25]
    return []


async def auto_caract(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
    # Récupère les deux cases précédentes
    type_choisi = interaction.namespace.type_commande
    nom_choisi = interaction.namespace.nom

    if type_choisi in BIBLIO_WE and nom_choisi in BIBLIO_WE[type_choisi]:
        choix = list(BIBLIO_WE[type_choisi][nom_choisi].keys())
        return [app_commands.Choice(name=c.capitalize(), value=c) for c in choix if current.lower() in c.lower()][:25]
    return []

async def auto_cmd_nom_craft(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
    choix = list(BIBLIO_CRAFT.keys())
    return [app_commands.Choice(name=c.capitalize(), value=c) for c in choix if current.lower() in c.lower()][:25]


@bot.tree.command(name="we", description="Affiche la commande World Edit pour une texture ou une forme")
@app_commands.describe(type_commande="Type de commande", nom="Nom de ce que l'on veut", caracteristique="Couleur ou variante")
@app_commands.autocomplete(type_commande=auto_type_cmd, nom=auto_nom, caracteristique=auto_caract)
async def commande_worldedit(interaction: discord.Interaction, type_commande: str, nom: str, caracteristique: str):
    type_propre = type_commande.lower().strip()
    nom_propre = nom.lower().strip()
    caract_propre = caracteristique.lower().strip()

    try:
        commandeWE = BIBLIO_WE[type_propre][nom_propre][caract_propre]
        # Vérification de la version
        if isinstance(commandeWE, dict):
            commande = commandeWE.get("we", "Erreur: Commande introuvable")
            lien_image = commandeWE.get("image", None)
        else:
            commande = commandeWE
            lien_image = None

        embed = discord.Embed(
            title=f"élément : {type_commande} -> {nom} -> {caracteristique}",
            color=discord.Color.teal()
        )
        embed.add_field(name=" ** </> ** Commande à copier", value=f"```\n{commande}\n```", inline=False)

        if lien_image:
            embed.set_image(url=lien_image)

        embed.set_footer(text="Clique sur la commande pour la copier")

        await interaction.response.send_message(embed=embed)
    except KeyError:
        await interaction.response.send_message("Impossible de trouver cette combinaison dans la base de données. Vérifie l'orthographe !",ephemeral=True)

@bot.tree.command(name="schematic", description="Envoie un schematic et ses coordonnées dans un channel discord")
@app_commands.describe(schematic="Fichier .schematic à envoyer", x = "Coordonnée X", y = "Coordonnée Y", z = "Coordonnée Z", image = "Jolie screen du build")
@app_commands.checks.has_role(1392508222358687824)
async def envoi_schematic(interaction: discord.Interaction, schematic: discord.Attachment, x: int, y: int, z:int, image: discord.Attachment):

    ID_channel = 1496971071985815722
    salon = await interaction.client.fetch_channel(ID_channel)


    if not schematic.filename.endswith(".schematic"):
        await interaction.response.send_message("❌ Erreur : Le fichier doit être un `.schematic` !", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    schem = await schematic.to_file()

    message = discord.Embed(
        title=f"<:Groupfqsf:1431359830676996167> Nouveau schematic disponible",
        description="Vous retrouvez ci-dessous les coordonnées ainsi que le fichier à télécharger !",
        color=discord.Color.green()
    )

    message.set_thumbnail(url=interaction.client.user.display_avatar.url)
    message.add_field(name="Coordonnées", value =f"X: **{x}** Y: **{y}** Z: **{z}**")
    message.add_field(name="Nom du schematic", value=f"**{schematic.filename}**")

    if not image.filename.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
        await salon.send(embed=message)
        await salon.send(file=schem)
        await interaction.edit_original_response(content="⚠️ L'image n'était pas au bon format. Envoi sans image.")
    else:
        fichier_image = await image.to_file()

        message.set_image(url=f"attachment://{fichier_image.filename}")
        await salon.send(embed=message, file=fichier_image)
        await salon.send(file=schem)
        await interaction.edit_original_response(content="✅ Le schematic et son image ont bien été envoyés !")


@envoi_schematic.error
async def envoi_schematic_erreur(interaction: discord.Interaction, error: app_commands.AppCommandError):

    if isinstance(error, app_commands.MissingRole):
        await interaction.response.send_message(
            "❌ Accès refusé : Tu n'as pas le rôle pour utiliser cette commande !", ephemeral=True)
    else:
        await interaction.response.send_message("⚠️ Une erreur inattendue s'est produite.", ephemeral=True)
        print(f"Erreur sur la commande schematic : {error}")

@bot.tree.command(name="craft", description="Calcul les ressources primaires pour le craft d'items compliqué")
@app_commands.describe(nom="Nom de l'item à décomposer", quantite="Quantité de l'item à craft")
@app_commands.autocomplete(nom=auto_cmd_nom_craft)
async def craft(interaction: discord.Interaction, nom: str, quantite: float):
    nom_propre = nom.lower().strip()
    quantite_float = float(quantite)

    quantite_propre = int(quantite_float) if quantite_float.is_integer() else quantite_float

    if nom_propre not in BIBLIO_CRAFT:
        await interaction.response.send_message(f"❌ Erreur : Je ne connais pas la recette pour `{nom_propre}`.", ephemeral=True)
        return
    await interaction.response.defer()

    lien_image = BIBLIO_CRAFT[nom_propre].get("image", None)

    res = recherche_craft(nom_propre, quantite, BIBLIO_CRAFT)


    texte_tmp = ""

    for matiere, qte in res.items():
        qte_items = math.ceil(qte)

        nb_dc = qte_items // 3456
        reste = qte_items % 3456

        nb_lignes = reste // 576
        reste = reste % 576

        nb_stacks = reste // 64
        reste_items = reste % 64

        parties = []
        if nb_dc > 0:
            parties.append(f"{nb_dc} DC")

        if nb_lignes > 0:
            parties.append(f"{nb_lignes} ligne{'s' if nb_lignes > 1 else ''}")

        if nb_stacks > 0:
            parties.append(f"{nb_stacks} stack{'s' if nb_stacks > 1 else ''}")

        if reste_items > 0 or not parties:
            parties.append(f"{reste_items}")

        if len(parties) > 1:
            texte_valeur = ", ".join(parties[:-1]) + " et " + parties[-1]
        else:
            texte_valeur = parties[0]

        texte_tmp += f"🔸 **{matiere.capitalize()}**: {texte_valeur}\n"

    embed = discord.Embed(
        title=f"Craft de {quantite_propre} {nom_propre}",
        description="Voici la liste des matières premières à récupérer : ",
        color=discord.Color.teal()
    )
    embed.add_field(name="Ressources totales nécessaires au craft\n", value=texte_tmp, inline=False)

    if lien_image:
        embed.set_image(url=lien_image)

    await interaction.followup.send(embed=embed)

if __name__ == "__main__":
    f = open('token.txt', 'r')
    bot.run(f.read().strip())
    f.close()